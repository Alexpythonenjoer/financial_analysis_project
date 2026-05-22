"""
Краулер: обход папки storage, извлечение текста из всех поддерживаемых форматов
(включая вложенные архивы) и сохранение результата в CSV.
"""

import os
import csv
import zipfile
import rarfile
import py7zr
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader

# Папка с файлами
STORAGE_DIR = './storage'
OUTPUT_CSV = './output/file_index.csv'
os.makedirs(os.path.dirname(OUTPUT_CSV), exist_ok=True)

def extract_text_from_docx(filepath):
    doc = Document(filepath)
    return '\n'.join([para.text for para in doc.paragraphs])

def extract_text_from_xlsx(filepath):
    wb = load_workbook(filepath, data_only=True)
    text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            text.append(' '.join([str(cell) for cell in row if cell is not None]))
    return '\n'.join(text)

def extract_text_from_pdf(filepath):
    reader = PdfReader(filepath)
    text = []
    for page in reader.pages:
        text.append(page.extract_text())
    return '\n'.join(text)

def extract_text_from_txt(filepath):
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()

# Расширение -> функция парсинга
PARSERS = {
    '.docx': extract_text_from_docx,
    '.xlsx': extract_text_from_xlsx,
    '.pdf': extract_text_from_pdf,
    '.txt': extract_text_from_txt,
}

def is_archive(filepath):
    """Проверяет, является ли файл архивом (zip, rar, 7z)"""
    ext = os.path.splitext(filepath)[1].lower()
    return ext in ['.zip', '.rar', '.7z']

def extract_archive(filepath, extract_to):
    """Распаковывает архив во временную папку и возвращает список извлечённых файлов"""
    os.makedirs(extract_to, exist_ok=True)
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.zip':
        with zipfile.ZipFile(filepath, 'r') as z:
            z.extractall(extract_to)
            return [os.path.join(extract_to, f) for f in z.namelist() if not f.endswith('/')]
    elif ext == '.rar':
        with rarfile.RarFile(filepath, 'r') as r:
            r.extractall(extract_to)
            return [os.path.join(extract_to, f) for f in r.namelist() if not f.endswith('/')]
    elif ext == '.7z':
        with py7zr.SevenZipFile(filepath, 'r') as sz:
            sz.extractall(extract_to)
            # py7zr не даёт простого списка, пройдёмся по extracted
            extracted = []
            for root, dirs, files in os.walk(extract_to):
                for f in files:
                    extracted.append(os.path.join(root, f))
            return extracted
    return []

def crawl_directory(root_dir):
    """Рекурсивно обходит root_dir, собирает все файлы и извлекает текст"""
    results = []  # (filepath, content)
    for dirpath, _, filenames in os.walk(root_dir):
        for filename in filenames:
            full_path = os.path.join(dirpath, filename)
            ext = os.path.splitext(filename)[1].lower()
            if ext in PARSERS:
                try:
                    content = PARSERS[ext](full_path)
                    results.append((full_path, content))
                except Exception as e:
                    print(f"Ошибка парсинга {full_path}: {e}")
            elif is_archive(full_path):
                # Распаковываем во временную папку
                temp_dir = os.path.join(root_dir, '_temp_extract_' + filename)
                try:
                    extracted_files = extract_archive(full_path, temp_dir)
                    for ef in extracted_files:
                        sub_ext = os.path.splitext(ef)[1].lower()
                        if sub_ext in PARSERS:
                            try:
                                content = PARSERS[sub_ext](ef)
                                results.append((ef, content))  # сохраняем путь к извлечённому файлу
                            except Exception as e:
                                print(f"Ошибка парсинга из архива {ef}: {e}")
                except Exception as e:
                    print(f"Ошибка распаковки {full_path}: {e}")
                finally:
                    # Удаляем временную папку
                    if os.path.exists(temp_dir):
                        import shutil
                        shutil.rmtree(temp_dir, ignore_errors=True)
    return results

def save_to_csv(data, csv_path):
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['file_path', 'content'])
        writer.writerows(data)

if __name__ == '__main__':
    print("Начинаем обход хранилища...")
    all_data = crawl_directory(STORAGE_DIR)
    print(f"Найдено {len(all_data)} файлов с текстом.")
    save_to_csv(all_data, OUTPUT_CSV)
    print(f"Результат сохранён в {OUTPUT_CSV}")